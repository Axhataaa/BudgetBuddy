"""
Settings -> "Export My Data": a complete, portable copy of everything
a user owns in BudgetBuddy - NOT the same feature as the Reports page
export (reports/services.py + utils/exportReport.js on the frontend),
which is a formatted financial report for a chosen date range. This
module has no date filtering at all - it always exports everything.

Security (see build_user_data_export()'s own docstring for the
per-field reasoning): every queryset here is filtered by
`user=<the authenticated user>`, and every serializer used is one
that's already used elsewhere in the API for that same user's own
data - none of them expose password hashes, JWT tokens, or any
other-user data, because none of them ever have.

Structure produced (matches the Settings feature spec exactly):

    BudgetBuddy_Data_Export_YYYY-MM-DD.zip
    |-- BudgetBuddy_Data.xlsx      (one sheet per category)
    |-- JSON/
    |   |-- profile.json
    |   |-- income.json
    |   |-- expenses.json
    |   |-- budgets.json
    |   |-- savings_goals.json
    |   |-- achievements.json
    |   |-- savings_transactions.json
    |   `-- notifications.json
    `-- README.txt

Performance (Section 14 of the spec): this runs synchronously inside
the request/response cycle today, same as every other view in this
project (no Celery/Redis here, matching the project's own standing
"no task queue" constraint). It's still structured so a background-job
migration is straightforward later: build_user_data_export(user) takes
only a user and returns plain bytes - no request/response object
anywhere inside it - so wrapping it in a task later
("write these bytes to storage, email a download link") wouldn't
require changing this function at all, only how its return value is
delivered.
"""

import io
import json
import zipfile
from datetime import datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from budgets.models import Budget, SavingsGoal, SavingsTransaction
from budgets.serializers import BudgetSerializer, SavingsGoalSerializer, SavingsTransactionSerializer
from expenses.models import Expense
from expenses.serializers import ExpenseSerializer
from incomes.models import Income
from incomes.serializers import IncomeSerializer
from notifications.models import Notification
from notifications.serializers import NotificationSerializer
from users.serializers import ProfileSerializer

HEADER_FONT = Font(bold=True)


def _autosize_columns(sheet, rows, headers):
    """Rough column-width heuristic (openpyxl doesn't auto-size) -
    widest of the header or any cell in that column, capped so one
    long description field can't blow out the whole sheet."""
    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, value in enumerate(row):
            widths[i] = min(max(widths[i], len(str(value))), 60)
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width + 2


def _write_table_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    for row in rows:
        sheet.append(row)
    _autosize_columns(sheet, rows, headers)


def _build_excel_workbook(
    profile_data, income_qs, expense_qs, budget_qs, goals_qs, achievements_qs, savings_transaction_qs, notification_qs
):
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default blank "Sheet"

    # --- Profile: key/value rows rather than a header row - it's one
    # record, not a table. ---
    profile_sheet = workbook.create_sheet(title="Profile")
    profile_sheet.append(["Field", "Value"])
    for cell in profile_sheet[1]:
        cell.font = HEADER_FONT
    for key, value in profile_data.items():
        if key == "profile_picture" and value is None:
            value = ""
        profile_sheet.append([key, "" if value is None else str(value)])
    _autosize_columns(profile_sheet, [[k, v] for k, v in profile_data.items()], ["Field", "Value"])

    _write_table_sheet(
        workbook,
        "Income",
        ["Date", "Source", "Amount", "Description"],
        [[i.date, i.source, float(i.amount), i.description] for i in income_qs],
    )

    _write_table_sheet(
        workbook,
        "Expenses",
        ["Date", "Title", "Category", "Amount", "Payment Method", "Description"],
        [
            [e.date, e.title, e.category, float(e.amount), e.payment_method, e.description]
            for e in expense_qs
        ],
    )

    _write_table_sheet(
        workbook,
        "Budgets",
        ["Category", "Month", "Year", "Monthly Limit"],
        [[b.category, b.month, b.year, float(b.monthly_limit)] for b in budget_qs],
    )

    _write_table_sheet(
        workbook,
        "Savings Goals",
        ["Goal Name", "Target Amount", "Current Amount", "Target Date", "Completed", "Purchased"],
        [
            [g.goal_name, float(g.target_amount), float(g.current_amount), g.target_date, g.is_completed, g.is_purchased]
            for g in goals_qs
        ],
    )

    _write_table_sheet(
        workbook,
        "Achievements",
        ["Goal Name", "Target Amount", "Purchase Date", "Purchase Note"],
        [
            [g.goal_name, float(g.target_amount), g.purchase_date, g.purchase_note]
            for g in achievements_qs
        ],
    )

    # Goal Name (denormalized from the FK) rather than a raw goal id -
    # matches every other sheet here, which shows the human-readable
    # label (category, source) rather than a database id.
    _write_table_sheet(
        workbook,
        "Savings Transactions",
        ["Goal Name", "Type", "Amount", "Note", "Date"],
        [
            [t.goal.goal_name, t.transaction_type, float(t.transaction_amount), t.note, t.created_at.strftime("%Y-%m-%d %H:%M")]
            for t in savings_transaction_qs
        ],
    )

    _write_table_sheet(
        workbook,
        "Notifications",
        ["Date", "Title", "Message", "Type", "Priority", "Read"],
        [
            [n.created_at.strftime("%Y-%m-%d %H:%M"), n.title, n.message, n.notification_type, n.priority, n.is_read]
            for n in notification_qs
        ],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_readme(generated_at, counts):
    return f"""BudgetBuddy Data Export
=======================

Export generated: {generated_at.strftime("%d %B %Y, %H:%M")} (UTC)

WHAT'S IN THIS ARCHIVE
-----------------------
This is a complete copy of the personal and financial data associated
with your BudgetBuddy account, provided in two equivalent formats:

  BudgetBuddy_Data.xlsx
      One spreadsheet with a separate sheet per category (Profile,
      Income, Expenses, Budgets, Savings Goals, Achievements,
      Savings Transactions, Notifications) - open this in Excel,
      Google Sheets, or any spreadsheet application.

  JSON/
      The same data, one file per category, in raw structured JSON -
      useful if you want to import this data into another tool or
      inspect it programmatically.

CONTENTS SUMMARY
-----------------
  Profile:              1 record
  Income records:       {counts["income"]}
  Expense records:      {counts["expenses"]}
  Budgets:               {counts["budgets"]}
  Savings goals:         {counts["savings_goals"]}
  Achievements:          {counts["achievements"]}
  Savings transactions:  {counts["savings_transactions"]}
  Notifications:         {counts["notifications"]}

WHAT'S NOT INCLUDED
--------------------
Your password, authentication tokens, and any other account security
credentials are intentionally EXCLUDED from this export. BudgetBuddy
never stores your password in a readable form, and it is never
included in any data export, for your security.

This export also never contains any other user's information - it is
scoped entirely to your own account.

Generated by BudgetBuddy.
"""


def build_user_data_export(user):
    """
    Returns the complete export as raw ZIP bytes for the given user.

    Every queryset below is explicitly filtered by `user=user` (or, for
    achievements, `user=user, is_archived=True` - the same filter the
    existing /budgets/savings-goals/achievements/ endpoint already
    uses) - there is no code path here that can return another user's
    data. Every serializer used (ProfileSerializer, IncomeSerializer,
    ExpenseSerializer, BudgetSerializer, SavingsGoalSerializer,
    SavingsTransactionSerializer, NotificationSerializer) is one
    already used elsewhere in the API for this exact same purpose
    (returning a user their own data), so none of them expose
    password hashes, JWT/refresh tokens, or any
    is_staff/is_superuser/admin-only field - those fields simply don't
    exist on Profile/Income/Expense/Budget/SavingsGoal/
    SavingsTransaction/Notification, and ProfileSerializer's own field
    list (users/serializers.py) never included them for the
    /users/me/ endpoint either. SavingsTransaction itself has no
    direct `user` field (see the query below), so it's scoped via
    `goal__user=user` instead - the same relationship
    SavingsTransactionViewSet already filters on.
    """
    generated_at = timezone.now()

    profile_data = ProfileSerializer(user.profile).data

    income_qs = Income.objects.filter(user=user).order_by("-date")
    expense_qs = Expense.objects.filter(user=user).order_by("-date")
    budget_qs = Budget.objects.filter(user=user).order_by("-year", "-month")
    goals_qs = SavingsGoal.objects.filter(user=user).order_by("-created_at")
    achievements_qs = SavingsGoal.objects.filter(user=user, is_archived=True).order_by("-purchase_date")
    # SavingsTransaction has no direct `user` FK (deposits/withdrawals
    # belong to a goal, and a goal belongs to a user) - scoped via
    # goal__user, the same relationship budgets/views.py's own
    # SavingsTransactionViewSet.get_queryset() already filters on for
    # exactly this reason. select_related("goal") avoids an N+1 query
    # per row when the Excel sheet below reads t.goal.goal_name.
    savings_transaction_qs = (
        SavingsTransaction.objects.filter(goal__user=user)
        .select_related("goal")
        .order_by("-created_at")
    )
    notification_qs = Notification.objects.filter(user=user).order_by("-created_at")

    json_files = {
        "profile.json": profile_data,
        "income.json": IncomeSerializer(income_qs, many=True).data,
        "expenses.json": ExpenseSerializer(expense_qs, many=True).data,
        "budgets.json": BudgetSerializer(budget_qs, many=True).data,
        "savings_goals.json": SavingsGoalSerializer(goals_qs, many=True).data,
        "achievements.json": SavingsGoalSerializer(achievements_qs, many=True).data,
        "savings_transactions.json": SavingsTransactionSerializer(savings_transaction_qs, many=True).data,
        "notifications.json": NotificationSerializer(notification_qs, many=True).data,
    }

    excel_bytes = _build_excel_workbook(
        profile_data, income_qs, expense_qs, budget_qs, goals_qs, achievements_qs, savings_transaction_qs, notification_qs
    )

    counts = {
        "income": income_qs.count(),
        "expenses": expense_qs.count(),
        "budgets": budget_qs.count(),
        "savings_goals": goals_qs.count(),
        "achievements": achievements_qs.count(),
        "savings_transactions": savings_transaction_qs.count(),
        "notifications": notification_qs.count(),
    }
    readme_text = _build_readme(generated_at, counts)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("BudgetBuddy_Data.xlsx", excel_bytes)
        for filename, payload in json_files.items():
            archive.writestr(f"JSON/{filename}", json.dumps(payload, indent=2, default=str))
        archive.writestr("README.txt", readme_text)

    return zip_buffer.getvalue()


def export_filename():
    return f"BudgetBuddy_Data_Export_{datetime.now().strftime('%Y-%m-%d')}.zip"
